import os
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

dir_f = "listy123"
files = []
for filename in os.listdir(dir_f):
        files.append(filename)

common_file = []

for file in files:
    f_name = dir_f + '/' + file
    # print(f_name)

    workbook = openpyxl.load_workbook(f_name)
    worksheet = workbook.active
    max_row = worksheet.max_row + 1
    # print(max_row)
    max_column = worksheet.max_column + 1
    # print(max_column)

    outside_array = []
    for row in range(1, max_row):
        inside_array = []
        for col in range(1, max_column):
            value = worksheet.cell(row=row, column=col).value
            if value is None:
                value = ''
            inside_array.append(value)
        outside_array.append(inside_array)
        # print(outside_array)
    common_file.append(outside_array)
print(common_file)

array1 = [arr[0] for arr in common_file[0]]
# print(array1)
array2 = [arr[1] for arr in common_file[1]]
# print(array2)
array3 = [arr[2] for arr in common_file[2]]
# print(array3)

len_max = len(array1)
if len(array2) > len_max:
    len_max = len(array2)
if len(array3) > len_max:
    len_max = len(array3)
print(len_max)

sheet2 = []
sheet3 = []
sheet4 = []

new_array = []
for i in range(0, len_max):
    try:
        part1 = array1[i]
    except Exception as error:
        part1 = ""
    try:
        part2 = array2[i]
    except Exception as error:
        part2 = ""
    try:
        part3 = array3[i]
    except Exception as error:
        part3 = ""

    new_array.append([part1, part2, part3])
    sheet2.append(part1)
    sheet3.append(part2)
    sheet4.append(part3)
# print(new_array)

for row in range(0, len(new_array)):
    for col in range(0, len(new_array[col])):
        worksheet.cell(row=row + 1, column=col + 1).value = new_array[row][col]

worksheet = workbook.active

wsheet2 = workbook.create_sheet("sheet2")
wsheet3 = workbook.create_sheet("sheet3")
wsheet4 = workbook.create_sheet("sheet4")

max_row1 = wsheet2.max_row
max_row2 = wsheet3.max_row
max_row3 = wsheet4.max_row
max_col1 = wsheet2.max_column
max_col2 = wsheet3.max_column
max_col3 = wsheet4.max_column

for row in range(0, len(sheet2)):
    col = 1
    wsheet2.cell(row=row + 1, column= col).value = sheet2[row]
for row in range(0, len(sheet3)):
    col = 1
    wsheet3.cell(row=row + 1, column= col).value = sheet3[row]
for row in range(0, len(sheet4)):
    col = 1
    wsheet4.cell(row=row + 1, column= col).value = sheet4[row]
workbook.save("hw_example.xlsx")
















