import openpyxl
from openpyxl.utils import get_column_letter

workbook = openpyxl.load_workbook("C:/Users/User/Documents/GitHub/python2/.idea/Лист Microsoft Excel.xlsx")
work_list = workbook.active

max_row = work_list.max_row
print(max_row)
max_column = work_list.max_column
print(max_column)


outside_array = []

for row in range(1, max_row + 1):
    inside_array = []


    for col in range(1, max_column + 1):
        value = work_list.cell(row=row, column=col).value
        if value is None:
            value = ''

        inside_array.append(value)

    outside_array.append(inside_array)

# print(outside_array)

def lenght_array(array):
    return len(array)

for row in range(0, lenght_array(outside_array)):


    for col in range(0, lenght_array(outside_array[row])):
        if get_column_letter(row + 1) == "A":
            letter = "B"
        else:
            letter = "A"
        work_list[f'{letter}{col + 4}'] = outside_array[row][col]

workbook.save('work_list.xlsx')






